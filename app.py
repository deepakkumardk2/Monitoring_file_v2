import io
import os
import re
import tempfile
from pathlib import Path
from datetime import datetime, date, timedelta

from flask import Flask, request, send_file, render_template_string, jsonify

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("openpyxl is not installed. Please run: pip install -r requirements.txt")

app = Flask(__name__)

MONTHS = {
    'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
    'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12,
}

NFMS_TEMPLATE_MAPPING = {
    'DAILY PROFILE': 'DAILY',
    'BLOCK PROFILE': 'BLOCK',
    'BILLING': 'BILLING',
}


def norm_text(value):
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip().upper()


def is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def parse_intish(value):
    if is_number(value):
        return int(value)
    if isinstance(value, str):
        cleaned = value.replace(',', '').strip()
        if re.fullmatch(r'-?\d+(\.0+)?', cleaned):
            return int(float(cleaned))
    return None


def as_int(value):
    return parse_intish(value)


def row_numeric_sum(row):
    total = 0
    found = False
    for cell in row:
        parsed = parse_intish(cell)
        if parsed is not None:
            total += parsed
            found = True
    return total if found else None


def find_year_from_workbook(wb):
    years = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, datetime):
                    years.append(cell.year)
                elif isinstance(cell, date):
                    years.append(cell.year)
    return max(years) if years else datetime.today().year


def parse_business_date_from_sheet_name(sheet_name, default_year):
    upper = sheet_name.upper()
    m = re.search(r'(\d{1,2})\s*[-_/ ]?\s*([A-Z]{3})', upper)
    if not m:
        return None
    day = int(m.group(1))
    mon = MONTHS.get(m.group(2)[:3])
    if not mon:
        return None
    return date(default_year, mon, day)


def latest_sheet_by_date(wb):
    year = find_year_from_workbook(wb)
    candidates = []
    for ws in wb.worksheets:
        d = parse_business_date_from_sheet_name(ws.title, year)
        if d:
            candidates.append((d, ws))
    if candidates:
        candidates.sort(key=lambda x: x[0])
        return candidates[-1][1], candidates[-1][0]
    return wb.worksheets[-1], None


def extract_prepaid_summary_totals(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    total_consumers = None
    converted_to_prepaid = None

    for row in ws.iter_rows(values_only=True):
        label = norm_text(row[1] if len(row) > 1 else None)
        if 'TOTAL NUMBER OF CONSUMERS' in label:
            total_consumers = row_numeric_sum(row[2:])
        elif 'NUMBER OF CONSUMERS CONVERTED TO PREPAID' in label:
            converted_to_prepaid = row_numeric_sum(row[2:])

    if total_consumers is None:
        raise ValueError(f"Could not find 'Total Number of Consumers' in {path}")
    if converted_to_prepaid is None:
        raise ValueError(f"Could not find 'Number Of Consumers Converted to Prepaid' in {path}")

    return {
        'total_consumers': total_consumers,
        'converted_to_prepaid': converted_to_prepaid,
    }


def parse_processed_log(path):
    wb = load_workbook(path, data_only=True)
    ws, report_date = latest_sheet_by_date(wb)
    rows = list(ws.iter_rows(values_only=True))

    regular_totals = {}
    nfms = {}
    jio = {}
    monthly_billing_counts = {}
    context = None
    last_regular_profile = None

    def header_text(row, idx):
        return norm_text(row[idx]) if len(row) > idx else ''

    for row in rows:
        c1 = header_text(row, 0)
        c2 = header_text(row, 1)
        c3 = header_text(row, 2)
        c6 = header_text(row, 5)

        if c1 == 'PROCESSED DATE' and c2 == 'BATCH NO' and c3 == 'PROFILE TYPE':
            context = 'regular'
            last_regular_profile = None
            continue
        if c1 == 'RTC_DTTM' and c3 == 'NFMS':
            context = 'nfms'
            continue
        if c1 == 'RTC_DTTM' and c3 == 'JIO DATA':
            context = 'jio'
            continue
        if c1 == 'RTC_DTTM' and c3 == 'MONTHLY BILLING':
            context = 'monthly'
            continue
        if c1 == 'RTC_DTTM' and c3 in {'BD', 'ACCT MATER', 'ACCT MASTER'}:
            context = None
            continue
        if not any(v is not None and str(v).strip() != '' for v in row):
            continue

        if context == 'regular':
            profile = c3
            if profile:
                last_regular_profile = profile
            if c6 == 'TOTAL' and last_regular_profile:
                regular_totals[last_regular_profile] = {
                    'raw': as_int(row[6]) if len(row) > 6 else None,
                    'processed': as_int(row[7]) if len(row) > 7 else None,
                    'exception': as_int(row[8]) if len(row) > 8 else None,
                }
            continue

        if context == 'nfms':
            if isinstance(row[0], (datetime, date)) and c3:
                nfms[c3] = {
                    'eligible': as_int(row[6]) if len(row) > 6 else None,
                    'meter_shared': as_int(row[7]) if len(row) > 7 else None,
                    'records_shared': as_int(row[8]) if len(row) > 8 else None,
                }
            continue

        if context == 'jio':
            if isinstance(row[0], (datetime, date)) and c3:
                jio[c3] = as_int(row[6]) if len(row) > 6 else None
            continue

        if context == 'monthly':
            if isinstance(row[0], (datetime, date)) and c3:
                monthly_billing_counts[c3] = as_int(row[6]) if len(row) > 6 else None
            continue

    if report_date is None:
        jio_dates = [row[0] for row in rows if isinstance(row[0], (datetime, date))]
        if jio_dates:
            report_date = max(d.date() if isinstance(d, datetime) else d for d in jio_dates)
        else:
            report_date = datetime.today().date()

    return {
        'sheet_name': ws.title,
        'report_date': report_date,
        'regular_totals': regular_totals,
        'nfms': nfms,
        'jio': jio,
        'monthly_billing_counts': monthly_billing_counts,
    }


def excel_date_equal(value, target_date):
    if isinstance(value, datetime):
        return value.date() == target_date
    if isinstance(value, date):
        return value == target_date
    return False


# def find_previous_cumulative(ws, report_date):
#     header_row = None
#     for r in range(1, ws.max_row + 1):
#         v = ws.cell(r, 1).value
#         if isinstance(v, str) and 'CUMMU DATA' in v.upper():
#             header_row = r
#             break
#     if header_row is None:
#         raise ValueError("Could not find cumulative section in ")

    exact_value = None
    last_numeric_date = None
    last_numeric_value = None
    next_row_candidate = None

    for r in range(header_row + 1, ws.max_row + 50):
        d = ws.cell(r, 1).value
        v = ws.cell(r, 2).value

        if d is None and v is None:
            if next_row_candidate is None:
                next_row_candidate = r
            continue

        if excel_date_equal(d, report_date) and is_number(v):
            exact_value = int(v)

        if isinstance(d, (datetime, date)) and is_number(v):
            last_numeric_date = d.date() if isinstance(d, datetime) else d
            last_numeric_value = int(v)
            next_row_candidate = r + 1

    base_value = exact_value if exact_value is not None else last_numeric_value
    if base_value is None:
        raise ValueError('Could not determine previous cumulative value from Sheet2')

    return {
        'header_row': header_row,
        'base_value': base_value,
        'last_numeric_date': last_numeric_date,
        'next_row': next_row_candidate or (header_row + 1),
    }


def find_or_create_date_row(ws, target_date, search_start_row):
    for r in range(search_start_row, ws.max_row + 50):
        val = ws.cell(r, 1).value
        if excel_date_equal(val, target_date):
            return r
        if val is None:
            ws.cell(r, 1).value = target_date
            return r

    r = ws.max_row + 1
    ws.cell(r, 1).value = target_date
    return r


def parse_time_value(value):
    for fmt in ('%I:%M %p', '%I:%M:%S %p', '%H:%M', '%H:%M:%S'):
        try:
            return datetime.strptime(value.strip(), fmt).time()
        except Exception:
            pass
    raise ValueError("Invalid time format. Use 12:09 AM or 00:09:00")


def parse_duration_text(value):
    value = value.strip()
    if re.fullmatch(r'\d{1,2}:\d{2}:\d{2}', value):
        return value
    raise ValueError("Invalid duration format. Use HH:MM:SS")


def secure_name(filename, default_name):
    if not filename:
        return default_name
    name = os.path.basename(filename).strip()
    name = re.sub(r'[^A-Za-z0-9._-]+', '_', name)
    return name or default_name


def format_mdy_nozero(value):
    return f"{value.month}/{value.day}/{value.year}"


def parse_flexible_date(value):
    value = value.strip()
    for fmt in ('%m/%d/%Y', '%m/%d/%y', '%m-%d-%Y', '%m-%d-%y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(value, fmt).date()
        except Exception:
            pass
    raise ValueError("Invalid date format. Use MM/DD/YYYY, MM-DD-YYYY, or YYYY-MM-DD")


def normalize_sms_id(value):
    return ''.join(ch for ch in str(value or '') if ch.isdigit())


def parse_sms_raw_text(raw_text):
    text = raw_text or ''
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    result = {
        'date': None,
        'pu_code': 'PU01',
        'total_count': None,
        'id_counts': {},
    }

    date_match = re.search(r'DATE\s*:?\s*-?\s*([0-9]{1,4}[/-][0-9]{1,2}[/-][0-9]{1,4})', text, re.IGNORECASE)
    if date_match:
        result['date'] = parse_flexible_date(date_match.group(1))

    total_match = re.search(r'\b(PU\d+)\b.*?\bTOTAL\b\s*([0-9,]+)', text, re.IGNORECASE | re.DOTALL)
    if total_match:
        result['pu_code'] = total_match.group(1).upper()
        result['total_count'] = int(total_match.group(2).replace(',', ''))
    else:
        pu_match = re.search(r'\b(PU\d+)\b', text, re.IGNORECASE)
        if pu_match:
            result['pu_code'] = pu_match.group(1).upper()

    for line in lines:
        id_match = re.search(r'(\d{16,20})', line)
        count_match = re.search(r'(\d[\d,]*)\s*$', line)
        if id_match and count_match:
            sms_id = id_match.group(1)
            count = int(count_match.group(1).replace(',', ''))
            result['id_counts'][sms_id] = count

    if result['total_count'] is None and result['id_counts']:
        result['total_count'] = sum(result['id_counts'].values())

    return result


def find_row_by_col_text(ws, col_idx, text, start_row=1, end_row=None, exact=False):
    target = norm_text(text)
    end_row = end_row or ws.max_row
    for row_no in range(start_row, end_row + 1):
        value = norm_text(ws.cell(row_no, col_idx).value)
        if (value == target) if exact else (target in value):
            return row_no
    return None


def build_output_filename(ws, report_date):
    cluster = secure_name(ws['C1'].value or 'Triveni', 'Triveni')
    return f"Job Monitoring UPPCL_{report_date.strftime('%d%m%Y')}__{cluster}.xlsm"


def fill_template(
    template_path,
    prepaid_summary_path,
    processed_log_path,
    mdm_path,
    billing_count,
    time_of_run,
    daily_billing_duration,
    sms_raw_text,
    sms_date_override=None,
    monthly_increment=None
):
    prepaid = extract_prepaid_summary_totals(prepaid_summary_path)
    mdm = extract_prepaid_summary_totals(mdm_path or prepaid_summary_path)
    processed = parse_processed_log(processed_log_path)

    report_date = processed['report_date']

    wb = load_workbook(template_path, keep_vba=True)
    ws1 = wb['Sheet1']
    ws2 = wb['Sheet2']

    monthly_default = sum(v for v in processed['monthly_billing_counts'].values() if isinstance(v, int))
    # cum_info = find_previous_cumulative(ws2, report_date)
    sms_payload = parse_sms_raw_text(sms_raw_text)

    if monthly_increment is None:
        monthly_increment = monthly_default if monthly_default else 0

    sms_date = sms_date_override or sms_payload['date'] or (report_date + timedelta(days=1))
    new_cumulative = cum_info['base_value'] + monthly_increment

    prepaid_header_row = find_row_by_col_text(ws1, 2, 'PREPAID BILLING DATE', exact=True)
    prepaid_value_row = prepaid_header_row + 1 if prepaid_header_row else 3
    ws1[f'B{prepaid_value_row}'] = report_date
    ws1[f'C{prepaid_value_row}'] = 'YES'
    ws1[f'D{prepaid_value_row}'] = prepaid['converted_to_prepaid']
    ws1[f'E{prepaid_value_row}'] = billing_count
    ws1[f'F{prepaid_value_row}'] = time_of_run
    ws1[f'G{prepaid_value_row}'] = daily_billing_duration

    regular = processed['regular_totals']
    processed_header_row = find_row_by_col_text(ws1, 2, 'Processed Date', exact=True) or 5
    processed_row_map = {
        'DAILY': processed_header_row + 1,
        'BILLING': processed_header_row + 2,
        'BLOCK': processed_header_row + 3,
    }
    for profile, row_no in processed_row_map.items():
        values = regular.get(profile, {})
        ws1[f'B{row_no}'] = report_date
        ws1[f'D{row_no}'] = values.get('raw')
        ws1[f'E{row_no}'] = values.get('processed')
        ws1[f'F{row_no}'] = values.get('exception')

    nfms = processed['nfms']
    nfms_header_row = find_row_by_col_text(ws1, 2, 'Meter Data(RTC) DATE', exact=True) or 11
    nfms_rows = {
        nfms_header_row + 1: 'DAILY PROFILE',
        nfms_header_row + 2: 'BLOCK PROFILE',
        nfms_header_row + 3: 'BILLING',
    }
    for row_no, template_label in nfms_rows.items():
        source_profile = NFMS_TEMPLATE_MAPPING[template_label]
        values = nfms.get(source_profile, {})
        ws1[f'B{row_no}'] = report_date
        ws1[f'D{row_no}'] = values.get('eligible')
        ws1[f'E{row_no}'] = values.get('meter_shared')
        ws1[f'F{row_no}'] = values.get('records_shared')

    sms_header_row = find_row_by_col_text(ws1, 2, 'SMS Sent to consumer', exact=True) or 15
    ws1[f'C{sms_header_row}'] = f'DATE:-{format_mdy_nozero(sms_date)}'
    total_count = sms_payload['total_count'] if sms_payload['total_count'] is not None else 0
    ws1[f'D{sms_header_row}'] = f'SMS Total Count: {total_count:,}'
    ws1[f'D{sms_header_row + 1}'] = sms_payload['pu_code'] or 'PU01'

    for row_no in range(sms_header_row + 2, ws1.max_row + 1):
        proposed_id = normalize_sms_id(ws1[f'C{row_no}'].value)
        if not proposed_id:
            if row_no > sms_header_row + 2:
                break
            continue
        ws1[f'D{row_no}'] = sms_payload['id_counts'].get(proposed_id, 0)

    monthly_header_row = find_row_by_col_text(ws1, 2, 'Monthly Billing DATE', exact=True) or 30
    monthly_value_row = monthly_header_row + 1
    ws1[f'B{monthly_value_row}'] = report_date
    ws1[f'C{monthly_value_row}'] = mdm['total_consumers']
    ws1[f'D{monthly_value_row}'] = new_cumulative

    next_day = report_date + timedelta(days=1)
    row_for_next_day = find_or_create_date_row(ws2, next_day, cum_info['header_row'] + 1)
    ws2.cell(row_for_next_day, 2).value = new_cumulative

    jio = processed['jio']
    jio_header_row = find_row_by_col_text(ws1, 2, 'JIO Data Sharing:DATE', exact=True) or 32
    jio_value_row = jio_header_row + 1
    ws1[f'B{jio_value_row}'] = report_date
    ws1[f'C{jio_value_row}'] = jio.get('BLOCK')
    ws1[f'D{jio_value_row}'] = jio.get('DAILY')
    ws1[f'E{jio_value_row}'] = jio.get('BILLING')

    date_cells = [
        f'B{prepaid_value_row}',
        *(f'B{row_no}' for row_no in processed_row_map.values()),
        *(f'B{row_no}' for row_no in nfms_rows.keys()),
        f'B{monthly_value_row}',
        f'B{jio_value_row}',
    ]
    for cell_ref in date_cells:
        ws1[cell_ref].number_format = 'd/m/yyyy'

    ws1[f'F{prepaid_value_row}'].number_format = 'h:mm:ss AM/PM'

    number_cells = [
        f'D{prepaid_value_row}', f'E{prepaid_value_row}',
        *(f'{col}{row_no}' for row_no in processed_row_map.values() for col in ['D', 'E', 'F']),
        *(f'{col}{row_no}' for row_no in nfms_rows.keys() for col in ['D', 'E', 'F']),
        f'C{monthly_value_row}', f'D{monthly_value_row}',
        f'C{jio_value_row}', f'D{jio_value_row}', f'E{jio_value_row}',
    ]
    for row_no in range(sms_header_row + 2, ws1.max_row + 1):
        proposed_id = normalize_sms_id(ws1[f'C{row_no}'].value)
        if not proposed_id:
            if row_no > sms_header_row + 2:
                break
            continue
        number_cells.append(f'D{row_no}')
    for cell_ref in number_cells:
        ws1[cell_ref].number_format = '#,##0'

    ws2.cell(row_for_next_day, 1).number_format = 'd/m/yyyy'
    ws2.cell(row_for_next_day, 2).number_format = '#,##0'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, {
        'latest_processed_sheet': processed['sheet_name'],
        'report_date': report_date.strftime('%d/%m/%Y'),
        'report_date_file': report_date.strftime('%d%m%Y'),
        'sms_date': sms_date.strftime('%d/%m/%Y'),
        'sms_total_count': total_count,
        'new_cumulative': new_cumulative,
        'download_name': build_output_filename(ws1, report_date),
    }


HTML_PAGE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Job Monitoring Auto Fill</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        body {
            font-family: Arial, sans-serif;
            background: #f4f7fb;
            margin: 0;
            padding: 30px;
        }
        .container {
            max-width: 900px;
            margin: auto;
            background: white;
            border-radius: 12px;
            padding: 24px;
            box-shadow: 0 8px 24px rgba(0,0,0,0.08);
        }
        h1 {
            margin-top: 0;
            color: #1f3b73;
        }
        .grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 16px;
        }
        .full {
            grid-column: 1 / -1;
        }
        label {
            display: block;
            margin-bottom: 6px;
            font-weight: bold;
        }
        input, button {
            width: 100%;
            box-sizing: border-box;
            padding: 10px 12px;
            border: 1px solid #ccd6e0;
            border-radius: 8px;
            font-size: 14px;
        }
        button {
            background: #2563eb;
            color: white;
            border: none;
            cursor: pointer;
            font-weight: bold;
        }
        button:hover {
            background: #1d4ed8;
        }
        .help {
            color: #555;
            font-size: 13px;
        }
        #msg {
            margin-top: 16px;
            white-space: pre-wrap;
        }
        .err {
            color: #b91c1c;
        }
        .ok {
            color: #166534;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Job Monitoring Auto Fill</h1>

        <form id="uploadForm" class="grid">
            <div class="full">
                <label>Template XLSM</label>
                <input type="file" name="template_file" accept=".xlsm" required>
            </div>

            <div>
                <label>Prepaid Summary XLSX</label>
                <input type="file" name="prepaid_summary_file" accept=".xlsx,.xlsm" required>
            </div>

            <div>
                <label>Processed Data Log XLSX</label>
                <input type="file" name="processed_log_file" accept=".xlsx,.xlsm" required>
            </div>

            <div>
                <label>MDM Eligible Count File (Optional)</label>
                <input type="file" name="mdm_file" accept=".xlsx,.xlsm">
            </div>

            <div>
                <label>Billing Count</label>
                <input type="number" name="billing_count" required>
            </div>

            <div>
                <label>Time of Run</label>
                <input type="text" name="time_of_run" value="12:09 AM" required>
            </div>

            <div>
                <label>Daily Billing Duration</label>
                <input type="text" name="daily_billing_duration" value="01:44:32" required>
            </div>

            <div>
                <label>SMS Date (Optional)</label>
                <input type="text" name="sms_date" placeholder="MM/DD/YYYY or leave blank to auto-detect">
            </div>

            <div class="full">
                <label>SMS Raw Data</label>
                <textarea name="sms_raw_text" rows="12" style="width:100%; box-sizing:border-box; padding:10px 12px; border:1px solid #ccd6e0; border-radius:8px; font-size:14px;" placeholder="Paste full SMS block here, e.g. DATE:-4/24/2026 ... PU01 TOTAL 638,646 ..."></textarea>
                <div class="help">You can paste the full SMS section text. The app will auto-pick SMS date, total count, PU code, and ID-wise counts.</div>
            </div>

            <div class="full">
                <label>Monthly Data Shared to Add (Optional)</label>
                <input type="number" name="monthly_increment" placeholder="Cumulative data shared for the month">
            </div>

            <div class="full">
                <button type="submit">Generate and Download</button>
            </div>
        </form>

        <div id="msg"></div>
    </div>

    <script>
        const form = document.getElementById('uploadForm');
        const msg = document.getElementById('msg');

        form.addEventListener('submit', async function(e) {
            e.preventDefault();
            msg.innerHTML = 'Processing... please wait.';

            try {
                const formData = new FormData(form);

                const response = await fetch('/generate', {
                    method: 'POST',
                    body: formData
                });

                if (!response.ok) {
                    let errorMessage = 'Failed to generate file';
                    try {
                        const err = await response.json();
                        errorMessage = err.error || errorMessage;
                    } catch (_) {}
                    throw new Error(errorMessage);
                }

                const blob = await response.blob();
                const disposition = response.headers.get('Content-Disposition') || '';
                let filename = 'output_filled.xlsm';
                const match = disposition.match(/filename="?([^"]+)"?/);
                if (match) filename = match[1];

                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = filename;
                document.body.appendChild(a);
                a.click();
                a.remove();
                window.URL.revokeObjectURL(url);

                msg.innerHTML = '<div class="ok">File generated successfully. Download started.</div>';
            } catch (error) {
                msg.innerHTML = '<div class="err">' + error.message + '</div>';
            }
        });
    </script>
</body>
</html>
"""


@app.route("/", methods=["GET"])
def index():
    return render_template_string(HTML_PAGE)


@app.route("/generate", methods=["POST"])
def generate():
    try:
        template_file = request.files.get("template_file")
        prepaid_summary_file = request.files.get("prepaid_summary_file")
        processed_log_file = request.files.get("processed_log_file")
        mdm_file = request.files.get("mdm_file")

        if not template_file or not template_file.filename:
            return jsonify({"error": "Template XLSM file is required."}), 400
        if not prepaid_summary_file or not prepaid_summary_file.filename:
            return jsonify({"error": "Prepaid Summary file is required."}), 400
        if not processed_log_file or not processed_log_file.filename:
            return jsonify({"error": "Processed Data Log file is required."}), 400

        billing_count_raw = request.form.get("billing_count", "").strip()
        time_of_run_raw = request.form.get("time_of_run", "").strip()
        daily_billing_duration_raw = request.form.get("daily_billing_duration", "").strip()
        sms_date_raw = request.form.get("sms_date", "").strip()
        sms_raw_text = request.form.get("sms_raw_text", "").strip()
        monthly_increment_raw = request.form.get("monthly_increment", "").strip()

        if not billing_count_raw:
            return jsonify({"error": "Billing Count is required."}), 400
        if not time_of_run_raw:
            return jsonify({"error": "Time of Run is required."}), 400
        if not daily_billing_duration_raw:
            return jsonify({"error": "Daily Billing Duration is required."}), 400
        if not sms_raw_text:
            return jsonify({"error": "SMS Raw Data is required."}), 400

        billing_count = int(billing_count_raw.replace(",", ""))
        time_of_run = parse_time_value(time_of_run_raw)
        daily_billing_duration = parse_duration_text(daily_billing_duration_raw)
        sms_date_override = parse_flexible_date(sms_date_raw) if sms_date_raw else None

        monthly_increment = None
        if monthly_increment_raw:
            monthly_increment = int(monthly_increment_raw.replace(",", ""))

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = os.path.join(tmpdir, secure_name(template_file.filename, "template.xlsm"))
            prepaid_path = os.path.join(tmpdir, secure_name(prepaid_summary_file.filename, "prepaid_summary.xlsx"))
            processed_path = os.path.join(tmpdir, secure_name(processed_log_file.filename, "processed_log.xlsx"))

            template_file.save(template_path)
            prepaid_summary_file.save(prepaid_path)
            processed_log_file.save(processed_path)

            mdm_path = prepaid_path
            if mdm_file and mdm_file.filename:
                mdm_path = os.path.join(tmpdir, secure_name(mdm_file.filename, "mdm.xlsx"))
                mdm_file.save(mdm_path)

            output_stream, meta = fill_template(
                template_path=template_path,
                prepaid_summary_path=prepaid_path,
                processed_log_path=processed_path,
                mdm_path=mdm_path,
                billing_count=billing_count,
                time_of_run=time_of_run,
                daily_billing_duration=daily_billing_duration,
                sms_raw_text=sms_raw_text,
                sms_date_override=sms_date_override,
                monthly_increment=monthly_increment
            )

            return send_file(
                output_stream,
                as_attachment=True,
                download_name=meta["download_name"],
                mimetype="application/vnd.ms-excel.sheet.macroEnabled.12"
            )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
